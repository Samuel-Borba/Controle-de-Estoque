import openpyxl
import datetime
import Banco_de_dados
import tkinter
from tkinter import filedialog
import tkinter.messagebox
import numpy as np
import PIL.Image
from PIL import Image, ImageDraw, ImageTk

class funcoes_geral():
    def str_para_data(self, mystr):
        if mystr.find('/') > -1:
            data_separada = mystr.split('/')
            for i in range(len(data_separada)):
                data_separada[i] = int(data_separada[i])
            data = datetime.date(
                data_separada[2],
                data_separada[1],
                data_separada[0]
            )
        else:
            data_separada = mystr.split('-')
            for i in range(len(data_separada)):
                data_separada[i] = int(data_separada[i])
            data = datetime.date(
                data_separada[0],
                data_separada[1],
                data_separada[2]
            )
        return data

    def str_para_data_e_hora(self, strData: str, strHora: str):
        hora_min = strHora.split(':')
        if strData.find('/') > -1:
            data_separada = strData.split('/')
            for i in range(len(data_separada)):
                data_separada[i] = int(data_separada[i])
            for i in range(len(hora_min)):
                hora_min[i] = int(hora_min[i])
            data_e_hora = datetime.datetime(
                data_separada[2],
                data_separada[1],
                data_separada[0],
                hora_min[0],
                hora_min[1]
            )
        else:
            data_separada = strData.split('-')
            for i in range(len(data_separada)):
                data_separada[i] = int(data_separada[i])
            for i in range(len(hora_min)):
                hora_min[i] = int(hora_min[i])
            data_e_hora = datetime.datetime(
                data_separada[0],
                data_separada[1],
                data_separada[2],
                hora_min[0],
                hora_min[1],
            )
        return data_e_hora

    def abrir_login(self, event):
        self.tela.destroy()
        self.chamar_login()

    def deletar_frames(self, tela, sit=''):
        id_frame = ''
        if sit != '':
            id_frame = sit
        for item in tela.winfo_children():
            if id_frame != '':
                if item.winfo_id() == id_frame: continue
            if item.winfo_id() in [self.frame_abas.winfo_id(), self.frame_login.winfo_id()]: continue
            for filho in item.winfo_children():
                try:
                    filho.destroy()
                except:
                    print(f'{filho.winfo_name()} é None!')
            try:
                item.destroy()
            except:
                print(f'{item.winfo_name()} é None!')

    def select_tree(self, event):
        try:
            inf = list(self.treev_principal.treeview.item((self.treev_principal.treeview.selection())).values())[0]
        except ValueError:
            inf = 'sem inf'
        self.nGerar = False
        if self.bt_acao_estoque.selecionado == True: self.nGerar = True
        self.botoes_frame_interacao(inf=inf)
        '''
            self.treev_principal.treeview.selection()[0] retorna o iid(muito importante, pois nele tem um padrao
            de reconhecer se é item filho ou pai, cada iid filho vem com o iid pai, ex: (iidfilho_iidpai)
            T[supondo que a item pai é o 2p e o indice é 7: 7_2p]
        '''
        if self.matricula.find('adm') > -1:
            lista = [self.bt_notificacoes, self.bt_acao_estoque, self.bt_dash, self.bt_compra_descarte,
                     self.bt_cadastro]
            i = 0
            for item in lista:
                if item.selecionado == True: break
                i += 1
            funcoes = {
                '0': self.evento_tree_not,
                '1': self.evento_tree_acao_estoque,
                '2': self.evento_tree_dash,
                '3': self.evento_tree_compra_descarte,
                '4': self.evento_tree_cadastro
            }
            funcoes[str(i)]()

    def filtro_geral(self, linha):
        if len(self.frame_fitro.winfo_children()) == 0:
            print('não existe objetos para usar o filtro')
            self.filtro = True
            return
        wb = self.abrir_BD()
        ws = wb[self.nome_aba_f]
        lista_cbb = []
        lista_entry = []
        lista_esc = []
        if 1 <= self.quant_f:
            lista_entry.append(self.entry_1)
            lista_cbb.append(self.cbb_1)
            lista_esc.append([self.esc_1_1, self.esc_1_2, self.esc_1_3])
        if 2 <= self.quant_f:
            lista_entry.append(self.entry_2)
            lista_cbb.append(self.cbb_2)
            lista_esc.append([self.esc_2_1, self.esc_2_2, self.esc_2_3])
        if 3 <= self.quant_f:
            lista_entry.append(self.entry_3)
            lista_cbb.append(self.cbb_3)
            lista_esc.append([self.esc_3_1, self.esc_3_2, self.esc_3_3])
        if 4 <= self.quant_f:
            lista_entry.append(self.entry_4)
            lista_cbb.append(self.cbb_4)
            lista_esc.append([self.esc_4_1, self.esc_4_2, self.esc_4_3])
        if 5 <= self.quant_f:
            lista_entry.append(self.entry_5)
            lista_cbb.append(self.cbb_5)
            lista_esc.append([self.esc_5_1, self.esc_5_2, self.esc_5_3])
        lista_verdade = []
        for i in range(self.quant_f):
            caso = 0
            pergunta = ''
            tipo = ''
            if lista_cbb[i].get() == '': continue
            if lista_entry[i].entry.get() == '':
                lista_verdade.append(True)
                continue
            if str(lista_entry[i].label_Class['text']).find('data') > -1:
                data = str(lista_entry[i].entry.get()).split('/')
                if len(data) != 3:
                    print('sua data é invalida, consideraremos sem este filtro!')
                    lista_verdade.append(True)
                    continue
                pergunta = datetime.date(int(data[2]), int(data[1]), int(data[0]))
                tipo = 'data'
            if str(lista_entry[i].label_Class['text']).find('flo') > -1 or \
                    str(lista_entry[i].label_Class['text']).find('int') > -1:
                try:
                    pergunta = float(lista_entry[i].entry.get())
                    tipo = 'float'
                except ValueError:
                    pergunta = int(lista_entry[i].entry.get())
                    tipo = 'int'
                except ValueError:
                    print('seu numero é invalido, conderaremos sem este filtro!')
                    lista_verdade.append(True)
                    continue
            if str(lista_entry[i].label_Class['text']).find('str') > -1:
                try:
                    pergunta = lista_entry[i].entry.get()
                    tipo = 'str'
                except:
                    print(f'erro inesperado no filtro da posição {i}, conderaremos sem este filtro!')
                    lista_verdade.append(True)
                    continue
            for col in range(1, 1_000):
                if ws.cell(row=1, column=col).value == None: break
                if ws.cell(row=1, column=col).value == lista_cbb[i].get():
                    if ws.cell(row=linha, column=col).value == None:
                        print(f'o valor da [row={linha}, column={col}] = None. Será considerada como falso no filtro!')
                        lista_verdade.append(False)
                        break
                    if tipo == 'str':
                        if str(ws.cell(row=linha, column=col).value).find(pergunta) > -1:
                            lista_verdade.append(True)
                            break
                        lista_verdade.append(False)
                        break
                    if tipo in ['float', 'int']:
                        for w in range(3):
                            if lista_esc[i][w]['bg'] == 'lightblue':
                                caso = w + 1
                                break
                        if caso == 1:
                            if float(ws.cell(row=linha, column=col).value) >= pergunta:
                                lista_verdade.append(True)
                                break
                            lista_verdade.append(False)
                            break
                        if caso == 2:
                            if float(ws.cell(row=linha, column=col).value) == pergunta:
                                lista_verdade.append(True)
                                break
                            lista_verdade.append(False)
                            break
                        if caso == 3:
                            if float(ws.cell(row=linha, column=col).value) <= pergunta:
                                lista_verdade.append(True)
                                break
                            lista_verdade.append(False)
                            break
                    if tipo == 'data':
                        data = datetime.datetime.strftime(ws.cell(row=linha, column=col).value, '%d/%m/%Y').split('/')
                        data = datetime.date(int(data[2]), int(data[1]), int(data[0]))
                        for w in range(3):
                            if lista_esc[i][w]['bg'] == 'lightblue':
                                caso = w + 1
                                break
                        if caso == 1:
                            if data >= pergunta:
                                lista_verdade.append(True)
                                break
                            lista_verdade.append(False)
                            break
                        if caso == 2:
                            if data == pergunta:
                                lista_verdade.append(True)
                                break
                            lista_verdade.append(False)
                            break
                        if caso == 3:
                            if data <= pergunta:
                                lista_verdade.append(True)
                                break
                            lista_verdade.append(False)
                            break
        for i in lista_verdade:
            if i == False:
                self.filtro = False
                return
        self.filtro = True

    def bind_cbb_filtro(self, event):
        wb = self.abrir_BD()
        ws = wb[self.nome_aba_f]
        tx = ''
        temp = None
        lista_cbb = []
        lista_entry = []
        lista_esc = []
        if 1 <= self.quant_f:
            lista_entry.append(self.entry_1)
            lista_cbb.append(self.cbb_1)
            lista_esc.append([self.esc_1_1, self.esc_1_2, self.esc_1_3])
        if 2 <= self.quant_f:
            lista_entry.append(self.entry_2)
            lista_cbb.append(self.cbb_2)
            lista_esc.append([self.esc_2_1, self.esc_2_2, self.esc_2_3])
        if 3 <= self.quant_f:
            lista_entry.append(self.entry_3)
            lista_cbb.append(self.cbb_3)
            lista_esc.append([self.esc_3_1, self.esc_3_2, self.esc_3_3])
        if 4 <= self.quant_f:
            lista_entry.append(self.entry_4)
            lista_cbb.append(self.cbb_4)
            lista_esc.append([self.esc_4_1, self.esc_4_2, self.esc_4_3])
        if 5 <= self.quant_f:
            lista_entry.append(self.entry_5)
            lista_cbb.append(self.cbb_5)
            lista_esc.append([self.esc_5_1, self.esc_5_2, self.esc_5_3])
        for i in range(self.quant_f):
            usando = False
            if lista_cbb[i].get() != '':
                for col in range(1, 1_000):
                    if ws.cell(row=1, column=col).value == None: break
                    if ws.cell(row=1, column=col).value == lista_cbb[i].get():
                        temp = ws.cell(row=2, column=col).value
                        tx = ''
                        if str(type(temp))[str(type(temp)).find("'") + 1:str(type(temp)).find("'") + 4] == 'dat':
                            tx = 'data'
                        else:
                            tx = str(type(temp))[str(type(temp)).find("'") + 1:str(type(temp)).find("'") + 4]
                        if lista_entry[i].label_Class['text'] != f'Filtro de {tx}': lista_entry[i].entry.delete(0,
                                                                                                                tkinter.END)
                        lista_entry[i].label_Class['text'] = f'Filtro de {tx}'
                        antigo_texto = lista_entry[i].entry.get()
                        if tx == 'str':
                            lista_entry[i].tipo_de_texto = 'texto'
                            lista_entry[i].chamada_tipo_de_texto()
                        if tx in ['flo', 'int']:
                            lista_entry[i].tipo_de_texto = 'numero'
                            lista_entry[i].chamada_tipo_de_texto()

                        if tx == 'data':
                            lista_entry[i].tipo_de_texto = 'data'
                            lista_entry[i].chamada_tipo_de_texto()
                        lista_entry[i].entry.insert(0,
                                                    antigo_texto)  # o metodo chamada_tipo_de_texto reseta a varString()
                        if tx in ['flo', 'int', 'data']:
                            for item in lista_esc[i]:
                                if item['fg'] == 'green':
                                    usando = True
                                    break
                            if usando == True: continue
                            for y in range(3):
                                lista_esc[i][y].configure(fg='black', bg='white', cursor='hand2')
                        else:
                            for y in range(3):
                                lista_esc[i][y].configure(fg='white', bg='white', cursor='arrow')
        self.fechar_BD(wb)

    def escolha_sinal_filtro(self, event):
        lista_esc = []
        posicao = 0
        achou = False
        if 1 <= self.quant_f:
            lista_esc.append([self.esc_1_1, self.esc_1_2, self.esc_1_3])
        if 2 <= self.quant_f:
            lista_esc.append([self.esc_2_1, self.esc_2_2, self.esc_2_3])
        if 3 <= self.quant_f:
            lista_esc.append([self.esc_3_1, self.esc_3_2, self.esc_3_3])
        if 4 <= self.quant_f:
            lista_esc.append([self.esc_4_1, self.esc_4_2, self.esc_4_3])
        if 5 <= self.quant_f:
            lista_esc.append([self.esc_5_1, self.esc_5_2, self.esc_5_3])
        if event.widget['cursor'] == 'hand2':
            if event.widget['fg'] == 'black':
                for i in range(self.quant_f):
                    if achou == True: break
                    for item in lista_esc[i]:
                        if item['fg'] == 'white': break
                        if item.winfo_id() == event.widget.winfo_id():
                            posicao = i
                            achou = True
                            break
                for x in range(3):
                    if lista_esc[posicao][x].winfo_id() != event.widget.winfo_id():
                        lista_esc[posicao][x].configure(fg='black', bg='white', cursor='hand2')
                        continue
                    event.widget.configure(fg='green', bg='lightblue', cursor='arrow')

    def pegar_imagem(self):  # Luhmeiy
        self.file_path = filedialog.askopenfilename(filetypes=[("Team 29", ".png .jpg .gif")])

    def salvar_imagem(self, matricula='', tamanho=(100, 100)):  # Luhmeiy
        if matricula == '': return
        if self.file_path in [None, '']: return
        novo_path = 'imagens\\banco-de-imagens\\'
        img = PIL.Image.open(self.file_path).convert("RGB")

        # Salva a imagem original
        img.save(f'{novo_path}{matricula}-original.png')

        def crop_center(pil_img):
            img_width, img_height = pil_img.size
            return pil_img.crop(((img_width - min(pil_img.size)) // 2,
                                 (img_height - min(pil_img.size)) // 2,
                                 (img_width + min(pil_img.size)) // 2,
                                 (img_height + min(pil_img.size)) // 2))

        # Edita e salva a imagem em retangulo
        img = crop_center(img)
        img = img.resize(tamanho, PIL.Image.Resampling.LANCZOS)
        img.save(f'{novo_path}{matricula}-retangular.png')  # retangular dimensionado
        npImage = np.array(img)
        h, w = img.size
        alpha = PIL.Image.new('L', img.size, 0)
        draw = ImageDraw.Draw(alpha)
        draw.pieslice([0, 0, h, w], 0, 360, fill=255)

        # Converte imagem para array
        npAlpha = np.array(alpha)
        npImage = np.dstack((npImage, npAlpha))
        PIL.Image.fromarray(npImage).save(f'{novo_path}{matricula}-circular.png')

    def carregar_tabela_filtro(self, *args):
        if self.caso_pai_f != '':
            self.carregar_tabela_com_codigo(nome_aba=self.nome_aba_f, caso_pai=self.caso_pai_f,
                                            colunas=self.colunas_f, treeview=self.treev_filtro,
                                            filtro=True, tags=self.tags_f)
            return
        self.carregar_tabela(nome_aba=self.nome_aba_f, colunas=self.colunas, treeview=self.treev_filtro, filtro=True,
                             tags=self.tags_f)

    def carregar_cabecalho_treeview(self, nome_aba, colunas: list):
        ''' as colunas é uma lista com o numero das colunas a serem adicionadas
        ex.: [0,1,5,7]'''
        wb = self.abrir_BD()
        ws = wb[nome_aba]
        # a variavel cabecalho vem com self porque queremos usar na classe app!
        self.cabecalho = []
        for col in colunas:
            if ws.cell(row=1, column=col).value == None: break
            self.cabecalho.append(ws.cell(row=1, column=col).value)
        self.fechar_BD(wb)

    def tags_personalizadas(self, linha: int, coluna_e_codigo: list, ws: openpyxl.Workbook):
        '''como deve ser escrito a lista colunas: colunas = [[numero coluna,texto-tag,texto-tag]...]
        exemplo: [[2,pendente-perigo,espera-alerta]] (esse caso irá usar a tag perigo para o texto pendente)'''
        tag = 'normal'
        if len(coluna_e_codigo) == 1:
            print(f'coluna {coluna_e_codigo[0]} sem informações para tags')
            return tag
        for i in range(1, len(coluna_e_codigo), 1):
            if ws.cell(row=linha, column=coluna_e_codigo[0]).value == coluna_e_codigo[i][
                                                                      :str(coluna_e_codigo[i]).find('-')]:
                tag = coluna_e_codigo[i][str(coluna_e_codigo[i]).find('-') + 1:]
                return tag
        return tag

    def carregar_tabela_com_codigo(self, nome_aba, caso_pai, colunas=False, treeview='', filtro=False, tags=False):
        '''
        --caso_pai:
        o caso pai é uma lista com dois termos: [numero coluna, texto da condicional]
        ex: [6,'pedido']
        na coluna 6, toda linha que tiver escrito pedido

        --tags:
        as tags são confusas. para usar as tags, precisa escrever uma lista assim:
        ex: tags = [[2,pendente-perigo,espera-alerta..],[4,pendente-perigo,espera-alerta..]].
        sempre com duas listas internas, uma é pra escolher a tag do item pai, outra dos filhos.
        dentro dessas listas estão os parametros: [numero coluna,texto-tag,texto-tag...]
        '''
        if tags != False:
            if len(tags) != 2:
                print('as tags desta função precisam de 2 listas no formato [num col, texto-tag,texto-tag..]')
                return
        if treeview == '': treeview = self.treev_principal
        if colunas == False: colunas = self.cabecalho
        treeview.treeview.delete(*treeview.treeview.get_children())
        wb = self.abrir_BD()
        ws = wb[nome_aba]
        linha = []
        # trazendo codigos unicos[coluna de codigos] para o loop
        lista_pai = []
        item_adicionado = False  # variavel de controle
        for lin in range(2, 1_000_000, 1):
            item_adicionado = False
            if ws.cell(row=lin, column=1).value == None: break
            if filtro == True:
                self.filtro_geral(linha=lin)
                if self.filtro == False: continue
            for item in lista_pai:
                if ws.cell(row=lin, column=1).value == item:
                    item_adicionado = True
                    break
            if item_adicionado == True: continue
            lista_pai.append(ws.cell(row=lin, column=1).value)
        indice = 0
        indice_Pai = 0
        for pathern in lista_pai:
            linha = []
            for lin in range(2, 1_000_000, 1):
                if ws.cell(row=lin, column=1).value == None: break
                if ws.cell(row=lin, column=1).value == pathern:
                    if ws.cell(row=lin, column=caso_pai[0]).value == caso_pai[1]:
                        minha_tag = self.tags_personalizadas(linha=lin, coluna_e_codigo=tags[0], ws=ws)
                        for x in colunas:
                            if x == 1: continue
                            linha.append(ws.cell(row=lin, column=x).value)
                        treeview.treeview.insert('', index=indice, iid=str(indice_Pai) + 'p',
                                                 text=ws.cell(row=lin, column=1).value, values=linha, tags=minha_tag)
                        indice += 1
                        break
            for lin in range(2, 1_000_000, 1):
                if ws.cell(row=lin, column=1).value == None: break
                if ws.cell(row=lin, column=1).value == pathern:
                    if ws.cell(row=lin, column=caso_pai[0]).value == caso_pai[1]: continue
                    minha_tag = self.tags_personalizadas(linha=lin, coluna_e_codigo=tags[1], ws=ws)
                    linha = [ws.cell(row=lin, column=x).value for x in colunas if x != 1]
                    treeview.treeview.insert(str(indice_Pai) + 'p', index=indice,
                                             iid=str(indice) + '_' + str(indice_Pai) + 'p', values=linha,
                                             tags=minha_tag)
                    indice += 1
            indice_Pai += 1
        self.fechar_BD(wb)

    def carregar_tabela(self, nome_aba, colunas: list, treeview='', filtro=False, tags=False):
        if tags != False:
            if len(tags) != 1:
                print('as tags desta função precisam somente de 1 lista no formato [num col, texto-tag,texto-tag..]')
                return
        if treeview == '': treeview = self.treev_principal
        treeview.treeview.delete(*treeview.treeview.get_children())
        wb = self.abrir_BD()
        ws = wb[nome_aba]
        linha = []
        indice = 0
        for lin in range(2, 1_000_000, 1):
            if ws.cell(row=lin, column=1).value == None: break
            if filtro == True:
                self.filtro_geral(lin)
                if self.filtro == False: continue
            minha_tag = self.tags_personalizadas(linha=lin, coluna_e_codigo=tags[0], ws=ws)
            linha = [ws.cell(row=lin, column=x).value for x in colunas]
            treeview.treeview.insert('', index=indice, iid=indice, values=linha, tags=minha_tag)
            indice += 1


class funcoes_login(Banco_de_dados.BD):
    def chamar_login(self, ws, row):
        self.matricula = ws.cell(row=row, column=1).value
        self.nome = ws.cell(row=row, column=4).value
        self.numero = ws.cell(row=row, column=5).value
        self.turno = ws.cell(row=row, column=6).value
        self.equipe = ws.cell(row=row, column=7).value

    def verificar_login(self, event):
        BD = self.abrir_BD()
        ws = BD['logins']

        for lin in range(2, 1_000_000, 1):
            if ws.cell(row=lin, column=1).value == None:
                tkinter.messagebox.showinfo('Login', 'Matricula não existe!')
                self.usuario.entry.delete(0, tkinter.END)
                self.senha.entry.delete(0, len(self.senha.entry.get()))
                break
            if ws.cell(row=lin, column=1).value == self.usuario.entry.get():
                if ws.cell(row=lin, column=2).value == self.senha.entry.get():
                    self.tela_login.destroy()
                    self.chamar_login(ws=ws, row=lin)
                    if str(ws.cell(row=lin, column=1).value).find('adm') > -1:
                        self.abrir_appAdm(matricula=self.matricula, nome=self.nome, numero=self.numero,
                                          turno=self.turno,
                                          equipe=self.equipe)
                    else:
                        self.abrir_appCli()
                else:
                    tkinter.messagebox.showinfo('Login', 'Senha errada!')
                    self.senha.entry.delete(0, len(self.senha.entry.get()))
                break
        self.fechar_BD(BD)


class funcoes_adm(Banco_de_dados.BD, funcoes_geral):
    def clicar_menu(self, event, clic='vazio'):
        self.nGerar = False
        lista = [self.bt_notificacoes, self.bt_acao_estoque, self.bt_dash, self.bt_compra_descarte,
                 self.bt_cadastro]
        if clic != 'vazio':
            item = lista[clic]
            for itemLista in lista:
                if itemLista.lbBase.winfo_name() == item.lbBase.winfo_name():
                    if itemLista.selecionado == True: break
                    itemLista.selecionou('')
                    break
            for itemLista in lista:
                if itemLista.lbBase.winfo_name() != item.lbBase.winfo_name():
                    itemLista.retira_selecao()
            pergunta = ''
            # if lista[clic].winfo_id() == lista[4].lbBase.winfo_id(): pergunta= self.frame_aba_cad.winfo_id()
            self.deletar_frames(tela=self.tela, sit=pergunta)
            self.gerar_aba_treeview()
            funcoes = {
                '0': self.gerar_obj_notificacoes,
                '1': self.gerar_obj_acao_estoque,
                '2': self.gerar_obj_dash,
                '3': self.gerar_obj_compra_descarte,
                '4': self.gerar_obj_cadastro
            }
            funcoes[str(clic)]()
            return
        contador = 0
        posicao = 0
        for item in self.frame_abas.winfo_children():
            if event.widget.winfo_id() == item.winfo_id():
                for itemLista in lista:
                    if itemLista.lbBase.winfo_name() == item.winfo_name():
                        if itemLista.selecionado == True: break
                        itemLista.selecionou('')
                        posicao = contador
                        break
                    contador += 1
                for itemLista in lista:
                    if itemLista.lbBase.winfo_name() != item.winfo_name():
                        itemLista.retira_selecao()
                break
        pergunta = ''
        # if event.widget.winfo_id() == lista[4].lbBase.winfo_id(): pergunta= self.frame_aba_cad.winfo_id()
        self.deletar_frames(tela=self.tela, sit=pergunta)
        self.gerar_aba_treeview()
        funcoes = {
            '0': self.gerar_obj_notificacoes,
            '1': self.gerar_obj_acao_estoque,
            '2': self.gerar_obj_dash,
            '3': self.gerar_obj_compra_descarte,
            '4': self.gerar_obj_cadastro
        }
        funcoes[str(posicao)]()

    def textos_ferramenta(self, id, nome_aba):
        wb = self.abrir_BD()
        ws = wb[nome_aba]
        nome = ''
        fabricante = ''
        part_number = ''
        for lin in range(1, 1_000_000):
            if ws.cell(row=lin, column=1).value == None: break
            if ws.cell(row=lin, column=1).value == id:
                nome = 'Nome: ' + str(ws.cell(row=lin, column=2).value)
                fabricante = 'Fabricante: ' + str(ws.cell(row=lin, column=4).value)
                part_number = 'PartNumber: ' + str(ws.cell(row=lin, column=6).value)
        self.fechar_BD(wb)
        return nome, fabricante, part_number

    def fechar_frame_to_BD(self, event):
        self.clicar_menu(event='', clic=0)
        self.frame_to_BD.destroy()

    def mostrar_entry_data(self, event):
        if event.widget.get() == 'aceito':
            self.entry_Data_liberacao.liberar_entry()
            self.entry_tempo_limite.liberar_entry()
            return
        self.entry_Data_liberacao.limpar_entry()
        self.entry_Data_liberacao.bloquear_entry()
        self.entry_tempo_limite.limpar_entry()
        self.entry_tempo_limite.bloquear_entry()

    def evento_add(self, event):
        self.bt_press = 'add'
        if self.verifica_se_respondido_notificacoes() == False:
            tkinter.messagebox.showinfo(title='Team-29', message='evento ja respondido!')
            return
        self.criar_frame_to_bd()
        lista = [self.bt_notificacoes, self.bt_acao_estoque, self.bt_dash, self.bt_compra_descarte,
                 self.bt_cadastro]
        i = 0
        for item in lista:
            if item.selecionado == True: break
            i += 1
        funcoes = {
            '0': self.bts_frameBD_not,
            '1': self.bts_frameBD_acao_estoque,
            '2': self.bts_frameBD_dash,
            '3': self.bts_frameBD_compra_descarte,
            '4': self.bts_frameBD_cadastro
        }
        funcoes[str(i)]()

    def enviarBD_notificacoes(self, event):
        if self.entry_Mensagem.entry.get() == '':
            tkinter.messagebox.showinfo(title='Team-29', message='Precisa preencher os campos!')
            return
        if self.cbb_conclusao.get() == '':
            tkinter.messagebox.showinfo(title='Team-29', message='Precisa preencher os campos!')
            return
        if self.cbb_conclusao.get() == 'aceito':
            if '' in [self.entry_Data_liberacao.entry.get(), self.entry_tempo_limite.entry.get()]:
                tkinter.messagebox.showinfo(title='Team-29', message='Precisa preencher os campos!')
                return
        wb = self.abrir_BD()
        ws = wb[self.nome_aba_f]
        sit = 'espera'
        if self.cbb_conclusao.get() in ['aceito', 'negado']: sit = 'respondido'
        data_str = self.data_pedido[:self.data_pedido.find(' ')]
        hora_str = self.data_pedido[self.data_pedido.find(' ') + 1:]
        data_ped = self.str_para_data_e_hora(str(data_str), str(hora_str))
        strDataNow = str(datetime.datetime.now())
        strDataNow = strDataNow[:strDataNow.find('.')]
        strDataNow_dat = strDataNow[:strDataNow.find(' ')]
        strDataNow_h = strDataNow[strDataNow.find(' ') + 1:]
        strDataNow = self.str_para_data_e_hora(strDataNow_dat, strDataNow_h)
        informacoes = []
        informacoes.append(self.label_cod['text'])
        informacoes.append(self.label_locutor['text'])
        informacoes.append('resposta')
        informacoes.append(self.id_ferramenta)
        informacoes.append(self.entry_Mensagem.entry.get())
        informacoes.append(sit)
        informacoes.append(self.cbb_conclusao.get())
        informacoes.append(data_ped)
        informacoes.append(strDataNow)

        if self.cbb_conclusao.get() == 'aceito':
            data_lib = self.str_para_data(self.entry_Data_liberacao.entry.get())
            informacoes.append(data_lib)
            informacoes.append(int(self.entry_tempo_limite.entry.get()))
        if self.cbb_conclusao.get() in ['aceito', 'negado']:
            for lin in range(1, 1_000_000):
                if ws.cell(row=lin, column=1).value == None:
                    tkinter.messagebox.showinfo('Team-29', 'não encontramos a pergunta no BD!')
                    self.fechar_BD(wb)
                    return
                if ws.cell(row=lin, column=1).value == self.codigo_selecionado:
                    ws.cell(row=lin, column=6).value = sit
                    ws.cell(row=lin, column=7).value = self.cbb_conclusao.get()
                    ws.cell(row=lin, column=9).value = strDataNow
                    if self.cbb_conclusao.get() == 'aceito':
                        ws.cell(row=lin, column=10).value = data_lib
                        ws.cell(row=lin, column=11).value = int(self.entry_tempo_limite.entry.get())
                    break
        for lin in range(1, 1_000_000):
            if ws.cell(row=lin, column=1).value == None:
                for i in range(1, len(informacoes) + 1):
                    ws.cell(row=lin, column=i).value = informacoes[i - 1]
                tkinter.messagebox.showinfo(title='Team-29', message='Resposta enviada!')
                self.fechar_BD(wb)
                self.fechar_frame_to_BD('')
                return
        self.fechar_BD(wb)

    def verifica_se_respondido_notificacoes(self):
        if self.respondido != 'espera': return False
        return True

    def escolher_aba_cadastro(self, event):
        if event != '':
            widget = event.widget
        else:
            widget = self.cbb_aba_cadastro
        if widget.get() == 'ferramentas':
            self.nome_aba = 'Tipos'
            self.colunas = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
            self.tags = [[6, ' -alerta']]  # por padrão, caso sem partNumb, colocar ' ' na inf
            self.deletar_frames(tela=self.tela, sit=self.frame_aba_cad.winfo_id())
            self.carregar_aba_cad_selecionada(nome_aba=self.nome_aba, colunas=self.colunas, tags=self.tags)
            return
        self.nome_aba = 'logins'
        self.colunas = [1, 3, 4, 5, 6, 7]
        self.tags = [[6, ' -alerta']]  # por padrão, caso sem partNumb, colocar ' ' na inf
        self.deletar_frames(tela=self.tela, sit=self.frame_aba_cad.winfo_id())
        self.carregar_aba_cad_selecionada(nome_aba=self.nome_aba, colunas=self.colunas, tags=self.tags)
    def enviarBD_ferramentas(self,event):
        nome_aba = 'Tipos'
        self.verificaId_ferramentas('')
        if self.idF == True:
            tkinter.messagebox.showinfo('Team-29', 'Esse id ja foi usado!')
            self.entry_codigo.entry.focus()
            return
        for item in self.frame_to_BD.winfo_children():
            if str(item).find('entry') > -1:
                if item.get() == '':
                    tkinter.messagebox.showinfo('Team-29', 'Preencha todos os campos!')
                    item.focus()
                    return
        if self.lb_caminho['text'] != 'S/N':
            self.salvar_imagem(matricula=str(self.entry_codigo.entry.get()))
        wb = self.abrir_BD()
        ws = wb['Tipos']
        uLinha = 1
        for lin in range(1,1_000_000):
            if ws.cell(row=lin,column=1).value in [None,'']:
                uLinha = lin
                break
        indice = 1
        lista = []
        for item in self.frame_to_BD.winfo_children():
            if str(item).find('entry') > -1:
                lista.append(str(item.get()))
        lista[5] = int(lista[5])
        if lista[6].find(','): lista[6] = lista[6].replace(',','.')
        lista[6] = float(lista[6])
        lista[10] = int(lista[10])
        for col in range(1,len(lista)+1):
            ws.cell(row=uLinha,column=col).value = lista[col-1]
        for item in self.frame_to_BD.winfo_children():
            if str(item).find('entry') > -1:
                item.delete(0,tkinter.END)
        self.lb_caminho['text'] = 'S/N'
        self.file_path = ''
        self.entry_codigo.entry.focus()
        self.fechar_BD(wb)
        tkinter.messagebox.showinfo('Team-29', f'Ferramenta[id: {lista[0]}, Nome: {lista[1]}] cadastrada!')
    def verificaId_ferramentas(self,event):
        nome_aba = 'Tipos'
        wb = self.abrir_BD()
        ws = wb[nome_aba]
        for lin in range(1,1_000_000):
            ID = ''
            if ws.cell(row=lin,column=1).value in [None,'']: break
            ID = str(ws.cell(row=lin,column=1).value)
            if self.entry_codigo.entry.get() == ID:
                if event != '': tkinter.messagebox.showinfo('Team-29', 'Esse id ja foi usado!')
                self.entry_codigo.entry.delete(0,tkinter.END)
                self.entry_codigo.entry.focus()
                self.fechar_BD(wb)
                self.idF = True
                return
        self.entry_codigo.verificaTexto('')
        self.fechar_BD(wb)
        self.idF = False
        return
    def pegar_imagemCadastroTec(self,event):
        self.pegar_imagem()
        self.lb_caminho['text']=self.file_path
    def nova_matricula(self, tipo='tecnico'):
        myVar = 'adm000'
        if tipo == 'tecnico': myVar='tec000'
        wb = self.abrir_BD()
        ws = wb['logins']
        indice = 1
        for lin in range(1,1_000_000):
            if ws.cell(row=lin,column=1).value in [None, '']: break
            if str(ws.cell(row=lin,column=1).value).find(myVar[:3]) > -1: indice +=1
        indice = str(indice)
        texto = myVar[:6-len(indice)] + indice
        return texto
    def verifica_senha(self, evento):
        senha1 = self.entry_senha
        senha2 = self.entry_senha2
        if evento != '':
            if senha2.entry.get() == '':
                senha2.verificaTexto('')
                return


        if senha1.entry.get() != senha2.entry.get():
            tkinter.messagebox.showinfo('Team-29', 'A senha de verificação está incorreta!')
            senha2.entry.delete(0,tkinter.END)
            senha2.entry.focus()
            return
        senha2.verificaTexto('')
    def enviarBD_tecnicos(self, event):
        if self.entry_senha.lb_forca['text'].find('fraca') > -1:
            tkinter.messagebox.showinfo('Team-29', 'Sua senha é fraca!')
            self.entry_senha.entry.delete(0,tkinter.END)
            self.entry_senha2.entry.delete(0, tkinter.END)
            self.entry_senha.verificaTexto('')
            self.entry_senha2.verificaTexto('')
            self.entry_senha.entry.focus()
            return

        for item in self.frame_to_BD.winfo_children():
            if str(item).find('entry') > -1:
                if item.get() == '':
                    tkinter.messagebox.showinfo('Team-29', 'Preencha todos os campos!')
                    item.focus()
                    return
        if self.lb_caminho['text'] != 'S/N':
            self.salvar_imagem(matricula=str(self.lb_matricula['text']))
        lista = []
        lista.append(str(self.lb_matricula['text']))
        lista.append(str(self.entry_senha.entry.get()))
        lista.append(str(self.entry_cpf.entry.get()))
        lista.append(str(self.entry_nome.entry.get()))
        lista.append(str(self.entry_telefone.entry.get()))
        lista.append(str(self.cbb_turno.get()))
        lista.append(str(self.entry_nomeEquipe.entry.get()))
        wb = self.abrir_BD()
        ws= wb['logins']
        uLinha = 0
        for i in range(1,1_000_000):
            if ws.cell(row=i,column=1).value in ['',None]:
                uLinha = i
                break
        for i in range(1,len(lista)+1):
            ws.cell(row=uLinha, column=i).value = str(lista[i-1])
        tkinter.messagebox.showinfo('Team-29', f'Tecnico[Matricula: {lista[0]}, Nome: {lista[3]}] cadastrado!')
        for item in self.frame_to_BD.winfo_children():
            if str(item).find('entry') > -1:
                item.delete(0,tkinter.END)
        self.lb_caminho['text'] = 'S/N'
        self.file_path = ''
        self.entry_senha.entry.focus()
        self.fechar_BD(wb)
    def carregar_nomeFerramentas(self):
        wb = self.abrir_BD()
        ws = wb['Tipos']
        self.lista_nomeFerramentas = []
        verificacao = False
        for lin in range(2,1_000_000):
            verificacao = False
            if ws.cell(row=lin,column=1).value in ['', None]: break
            for item in self.lista_nomeFerramentas:
                if item == str(ws.cell(row=lin,column=2).value).lower().strip():
                    verificacao = True
                    break
            if verificacao == False:
                    self.lista_nomeFerramentas.append(str(ws.cell(row=lin,column=2).value).lower().strip())
        self.cbb_nomeFerramenta['values'] = self.lista_nomeFerramentas
        self.fechar_BD(wb)
    def carregar_cbb_fabricantes(self, event):
        if event.widget.get() == '': return
        wb = self.abrir_BD()
        ws = wb['Tipos']
        self.cbb_fabricantes['state'] = 'default'
        self.cbb_id['state'] = 'default'
        self.cbb_fabricantes.delete(0, tkinter.END)
        self.cbb_id.delete(0, tkinter.END)
        self.cbb_fabricantes['state'] = 'readonly'
        self.cbb_id['state'] = 'readonly'
        for item in self.frame_to_BD.winfo_children():
            if str(item).lower().find('combobox') == -1:
                item.destroy()
        self.gerar_cab_frameBD_compra_descarte()
        self.lista_fabricantes = []
        filtro = event.widget.get()
        verificacao = False
        for lin in range(2,1_000_000):
            verificacao = False
            if ws.cell(row=lin,column=1).value in ['', None]: break
            if str(ws.cell(row=lin,column=2).value).lower().strip() == filtro:
                for item in self.lista_fabricantes:
                    if item == str(ws.cell(row=lin, column=4).value).lower().strip().title():
                        verificacao = True
                        break
                if verificacao == False:
                    self.lista_fabricantes.append(str(ws.cell(row=lin, column=4).value).lower().strip().title())
        self.cbb_fabricantes['values']=self.lista_fabricantes
        self.fechar_BD(wb)
    def carregar_cbb_id(self, event):
        if event.widget.get() == '': return
        wb = self.abrir_BD()
        ws = wb['Tipos']
        self.cbb_id['state'] = 'default'
        self.cbb_id.delete(0, tkinter.END)
        self.cbb_id['state'] = 'readonly'
        filtro_nome = self.cbb_nomeFerramenta.get()
        filtro_fab = event.widget.get()
        verificacao = False
        self.lista_id = []
        for lin in range(2,1_000_000):
            verificacao = False
            if ws.cell(row=lin,column=1).value in ['', None]: break
            if str(ws.cell(row=lin, column=2).value).lower().strip() == filtro_nome:
                if str(ws.cell(row=lin, column=4).value).lower().strip().title() == filtro_fab:
                    for item in self.lista_id:
                        if item == str(ws.cell(row=lin, column=1).value):
                            verificacao = True
                            break
                    if verificacao == False:
                        self.lista_id.append(str(ws.cell(row=lin, column=1).value))
        self.cbb_id['values'] = self.lista_id
        self.fechar_BD(wb)
    def verificar_cbbs_compra_descarte(self, event):
        if '' in [self.cbb_id.get(), self.cbb_fabricantes.get(), self.cbb_nomeFerramenta.get()]: return
        wb = self.abrir_BD()
        ws = wb['Tipos']
        for item in self.frame_to_BD.winfo_children():
            if str(item).lower().find('combobox') == -1:
                item.destroy()
        self.gerar_cab_frameBD_compra_descarte()
        linha = 0
        self.inf_ferramenta = []
        for i in range(2,1_000_000):
            if ws.cell(row=i, column=1).value in [None,'']: break
            if str(ws.cell(row=i,column=1).value) == str(self.cbb_id.get()):
                linha = i
                break
        for col in range(1,1_000):
            if ws.cell(row=1, column=col).value in [None,'']: break
            self.inf_ferramenta.append(ws.cell(row=linha, column=col).value)
        self.gerar_componentes_compra_descarte()
    def enviarBD_Compra(self, event):
        lista = [self.entry_quantidade, self.entry_valorUn, self.entry_data_compra]
        for item in lista:
            if item.entry.get() == '':
                tkinter.messagebox.showinfo('Team-29', 'O único campo não obrigatório é o campo [Mensagem]!')
                item.entry.focus()
                return
        data = self.str_para_data(self.entry_data_compra.entry.get())
        if data < datetime.date(1996,12,30):
            tkinter.messagebox.showinfo('Team-29', 'Não aceita datas anteriores a 1997!')
            self.entry_data_compra.entry.delete(0,tkinter.END)
            self.entry_data_compra.entry.focus()
            return
        if int(self.entry_quantidade.entry.get()) < 1:
            tkinter.messagebox.showinfo('Team-29', 'Não aceita quantidade menor que 1!')
            self.entry_quantidade.entry.delete(0, tkinter.END)
            self.entry_quantidade.entry.focus()
            return
        wb = self.abrir_BD()
        ws=wb['acoes estoque']
        quantidade = int(self.entry_quantidade.entry.get())
        valor_temp = str(self.entry_valorUn.entry.get())
        if valor_temp.find(','): valor_temp = valor_temp.replace(',','.')
        valor = float(valor_temp)
        login = self.matricula
        situacao = 'Nova'
        mensagem = self.entry_mensagem.entry.get()
        if self.entry_mensagem.entry.get() == '':
            mensagem = '-'
        acao = 'compra'
        id = str(self.cbb_id.get())
        uLinha = 0
        quant_pecas = 0
        for lin in range(1,1_000_000):
            if ws.cell(row=lin,column=1).value in [None,'']:
                uLinha = lin
                break
            if str(ws.cell(row=lin,column=1).value)[:len(id)] == id:
                if str(ws.cell(row=lin,column=6).value) == acao:
                    quant_pecas +=1
                    continue
        id += '-0000'
        for i in range(1,quantidade+1):
            n = str(quant_pecas+i)
            ws.cell(row=uLinha,column=1).value = id[:len(id)-len(n)] + n
            ws.cell(row=uLinha, column=2).value = login
            ws.cell(row=uLinha, column=3).value = valor
            ws.cell(row=uLinha, column=4).value = situacao
            ws.cell(row=uLinha, column=5).value = mensagem
            ws.cell(row=uLinha, column=6).value = acao
            ws.cell(row=uLinha, column=7).value = data
            uLinha +=1
        tkinter.messagebox.showinfo('Team-29', f'{quantidade} compras cadastradas no estoque!')
        self.frame_to_BD.destroy()
        self.bts_frameBD_compra_descarte()
        self.fechar_BD(wb)